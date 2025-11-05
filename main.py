import cv2
import numpy as np
from openpyxl.workbook import Workbook
from openpyxl.styles import Font

import helper
import config

import tkinter as tk
from tkinter import filedialog
from PIL import Image, ImageTk
import os

id=[]
res=[]
def processImages(filepath):


    img = cv2.imread(filepath)
    filename = os.path.basename(filepath)
    print("Filename:", filename)
    filename_without_extension = filename.split('.')[0]
    print("Filename without extension:", filename_without_extension)
    id.append(filename_without_extension)

   
    width = config.IMAGE_WIDTH
    height = config.IMAGE_HEIGHT
    questions = config.NUM_QUESTIONS
    choices = config.NUM_CHOICES
    ans = config.ANSWER_KEY


    img = cv2.resize(img, (width, height))
    cv2.imshow("Input image", img)
    cv2.waitKey(0)

    imgGray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    cv2.imshow("Grayscaled image", imgGray)
    cv2.waitKey(0)


    imgBlur = cv2.GaussianBlur(imgGray, config.BLUR_KERNEL_SIZE, config.BLUR_SIGMA)
    cv2.imshow("Blurred image", imgBlur)
    cv2.waitKey(0)


    imgCanny = cv2.Canny(imgBlur, config.CANNY_THRESHOLD1, config.CANNY_THRESHOLD2)
    cv2.imshow("Edge detected image", imgCanny)
    cv2.waitKey(0)

   #custom “connected-component contour tracing” algorithm.
    contours = helper.get_edge_points(imgCanny)
    print("No. of contours found:")
    print(len(contours)) 
    #it will work now as canvas
    drawCnt = img.copy()
    for i in range(len(contours)):

        helper.manual_draw_contours(drawCnt, contours[i], (0, 255, 0), 1)
       
    cv2.imshow("Contours", drawCnt)
    cv2.waitKey(0)





    corner_points = []
    for contour in contours:
        corner_list = helper.find_corners(contour)
        corner_points.append(corner_list)


    drawCorners = img.copy()
    for corner in corner_points:
        helper.manual_draw_contours(drawCorners, corner, (0, 255, 0), 2)
    cv2.imshow("Corners",drawCorners)
    cv2.waitKey(0)

    # FIND AREA
    # Corners are in (row, col) format
    areas = []
    for corners in corner_points:
        # corners = [(top_left_row, top_left_col), (top_right_row, top_right_col), 
        #            (bottom_left_row, bottom_left_col), (bottom_right_row, bottom_right_col)]
        
        # Width = difference in columns
        rec_width1 = abs(corners[1][1] - corners[0][1])  # top: right_col - left_col
        rec_width2 = abs(corners[3][1] - corners[2][1])  # bottom: right_col - left_col
        rec_width = (rec_width2 + rec_width1) / 2
        
        # Height = difference in rows
        rec_height1 = abs(corners[2][0] - corners[0][0])  # left: bottom_row - top_row
        rec_height2 = abs(corners[3][0] - corners[1][0])  # right: bottom_row - top_row
        rec_height = (rec_height2 + rec_height1) / 2
        
        area = rec_width * rec_height
        areas.append(area)

    sorted_area = sorted(areas, reverse=True)
    print(sorted_area)
    max_area = sorted_area[0]
    area_index = []
    for i in range(len(sorted_area)):
        for j in range(len(areas)):
            if (sorted_area[i] == areas[j]):
                area_index.append(j)
                break;
    # print(area_index)
    max_index = area_index[0]
    max_contour = img.copy()
    # 0 -> ans, 2 -> grade, 4 -> name
    helper.manual_draw_contours(max_contour, contours[area_index[0]], (0, 255, 0), 1)

    cv2.imshow("Answer section", max_contour)
    cv2.waitKey(0)
    
    
    

    ans_corner_points = corner_points[max_index]
    
    # Debug: Print corner points to verify
    print(f"\n📍 Answer Section Corner Points (row, col format):")
    print(f"   Top-Left:     {ans_corner_points[0]}")
    print(f"   Top-Right:    {ans_corner_points[1]}")
    print(f"   Bottom-Left:  {ans_corner_points[2]}")
    print(f"   Bottom-Right: {ans_corner_points[3]}")
    
    # Corners are in (row, col) format from find_corners()
    # Extract: top_left (index 0) and bottom_right (index 3)
    tl_row = ans_corner_points[0][0] + config.ANSWER_CORNER_OFFSET['top_left_y']   # row = y
    tl_col = ans_corner_points[0][1] + config.ANSWER_CORNER_OFFSET['top_left_x']   # col = x
    br_row = ans_corner_points[3][0] + config.ANSWER_CORNER_OFFSET['bottom_right_y']  # row = y
    br_col = ans_corner_points[3][1] + config.ANSWER_CORNER_OFFSET['bottom_right_x']  # col = x
    
    # Convert to integers and ensure they're within image bounds
    img_h, img_w = imgGray.shape[:2]
    tl_row = max(0, int(tl_row))
    tl_col = max(0, int(tl_col))
    br_row = min(img_h, int(br_row))
    br_col = min(img_w, int(br_col))
    
    print(f"\n📐 ROI Extraction:")
    print(f"   Top-Left (row, col):     ({tl_row}, {tl_col})")
    print(f"   Bottom-Right (row, col): ({br_row}, {br_col})")
    print(f"   Dimensions (h × w):      {br_row - tl_row} × {br_col - tl_col}")
    
    # Calculate width and height (always positive)
    wd = br_col - tl_col  # width in columns
    ht = br_row - tl_row  # height in rows
    
    # Validate dimensions
    if wd <= 0 or ht <= 0:
        print(f"⚠️ Warning: Invalid ROI dimensions (w={wd}, h={ht}). Using full contour area.")
        # Fallback: use original corner points without offset
        tl_row = int(ans_corner_points[0][0])
        tl_col = int(ans_corner_points[0][1])
        br_row = int(ans_corner_points[3][0])
        br_col = int(ans_corner_points[3][1])
        wd = br_col - tl_col
        ht = br_row - tl_row
    
    # For array slicing: imgGray[rows, cols] = imgGray[y, x]
    y, x, h, w = tl_row, tl_col, ht, wd
    roi = imgGray[y:y + h, x:x + w]
    
    # Validate ROI was extracted successfully
    if roi.size == 0:
        print(f"❌ Error: ROI extraction failed. Coordinates: ({x},{y}) size ({w}x{h})")
        print(f"   Image size: {imgGray.shape}")
        print(f"   Corner points: {ans_corner_points}")
        raise ValueError("Failed to extract valid ROI from answer section")
    
    ans_new_image = np.zeros_like(roi)  # Create a black image with the same size as roi
    ans_new_image[:, :] = roi
    
    # Use the full ROI directly - the corner detection + offsets already define the correct region
    # The smart auto-detection below can miss empty bubbles and crop the lower portion
    img_padded = ans_new_image
    
    # Optional: Apply smart filtering only if EDGE_EXCLUSION_MARGIN is enabled
    if config.EDGE_EXCLUSION_MARGIN > 0:
        # Smart ROI detection that excludes corner markers and edge noise
        # Apply threshold to find marked regions
        _, binary = cv2.threshold(ans_new_image, config.THRESHOLD_VALUE, 255, cv2.THRESH_BINARY_INV)
        
        # Find contours of all detected shapes
        contours_roi, _ = cv2.findContours(binary, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        
        if len(contours_roi) > 0:
            # Filter out corner markers by analyzing contour positions
            height, width = ans_new_image.shape[:2]
            
            # Define edge regions (corner markers are typically in outer portion of image)
            edge_margin = config.EDGE_EXCLUSION_MARGIN
            inner_x_start = int(width * edge_margin)
            inner_x_end = int(width * (1 - edge_margin))
            inner_y_start = int(height * edge_margin)
            inner_y_end = int(height * (1 - edge_margin))
            
            # Collect only contours in the central region (answer bubbles)
            central_contours = []
            for cnt in contours_roi:
                x_cnt, y_cnt, w_cnt, h_cnt = cv2.boundingRect(cnt)
                cx = x_cnt + w_cnt // 2  # Center x
                cy = y_cnt + h_cnt // 2  # Center y
                
                # Keep only if center is in the inner region
                if inner_x_start <= cx <= inner_x_end and inner_y_start <= cy <= inner_y_end:
                    central_contours.append(cnt)
            
            if len(central_contours) > 0:
                # Get bounding box of only the central contours (answer grid)
                all_points = np.vstack(central_contours)
                content_x, content_y, content_w, content_h = cv2.boundingRect(all_points)
                
                # Add configurable margin
                margin_top = config.BORDER_PADDING['top']
                margin_bottom = config.BORDER_PADDING['bottom']
                margin_left = config.BORDER_PADDING['left']
                margin_right = config.BORDER_PADDING['right']
                
                # Ensure margins don't go out of bounds
                y_start = max(0, content_y - margin_top)
                y_end = min(height, content_y + content_h + margin_bottom)
                x_start = max(0, content_x - margin_left)
                x_end = min(width, content_x + content_w + margin_right)
                
                # Extract the exact answer grid region
                img_padded = ans_new_image[y_start:y_end, x_start:x_end]
                print(f"✓ Smart ROI applied: {x_start}:{x_end}, {y_start}:{y_end}")
    
    cv2.imshow("Birds eye view", img_padded)
   
    cv2.waitKey(0)

    # Optimized thresholding - no need to pre-create imgThres
    imgThres = helper.thresholdImage(img_padded, None, config.THRESHOLD_VALUE)
    cv2.imshow("Thresholded image", imgThres)
    cv2.waitKey(0)

    boxes = helper.splitBoxes(imgThres)
    myPixelVal = np.zeros((questions, choices))
    countC = 0
    countR = 0

    for image in boxes:
        totalPixels = helper.countNonZeroPixel(image)

        myPixelVal[countR][countC] = totalPixels
        countC += 1
        if (countC == choices):
            countR += 1
            countC = 0

    myIndex = []
    for x in range(0, questions):
        arr = myPixelVal[x]
        myIndexVal = np.where(arr == np.amax(arr))
        # print(myIndexVal[0])
        myIndex.append(myIndexVal[0][0])

    gradings = []
    for x in range(0, questions):
        if (ans[x] == myIndex[x]):
            gradings.append(1)
        else:
            gradings.append(0)
    score = (sum(gradings) / questions) * 100
    res.append(score)

    imgResult = img_padded.copy()
    imgResult = helper.showAnswers(imgResult, myIndex, questions, ans, choices, gradings)
    cv2.imshow("Answers", imgResult)
    cv2.waitKey(0)


    grade_index = area_index[2]
    grade_contour = img.copy()


    grade_corner_points = corner_points[grade_index]

    # Corners are in (row, col) format from find_corners()
    grade_tl_row = grade_corner_points[0][0] + config.GRADE_CORNER_OFFSET['top_left_y']
    grade_tl_col = grade_corner_points[0][1] + config.GRADE_CORNER_OFFSET['top_left_x']
    grade_br_row = grade_corner_points[3][0] + config.GRADE_CORNER_OFFSET['bottom_right_y']
    grade_br_col = grade_corner_points[3][1] + config.GRADE_CORNER_OFFSET['bottom_right_x']
    
    # Convert to integers and ensure they're within image bounds
    img_h, img_w = imgGray.shape[:2]
    grade_tl_row = max(0, int(grade_tl_row))
    grade_tl_col = max(0, int(grade_tl_col))
    grade_br_row = min(img_h, int(grade_br_row))
    grade_br_col = min(img_w, int(grade_br_col))
    
    # Calculate width and height (always positive)
    grade_wd = grade_br_col - grade_tl_col
    grade_ht = grade_br_row - grade_tl_row
    
    # Validate dimensions
    if grade_wd <= 0 or grade_ht <= 0:
        print(f"⚠️ Warning: Invalid grade ROI dimensions (w={grade_wd}, h={grade_ht})")
        grade_tl_row = int(grade_corner_points[0][0])
        grade_tl_col = int(grade_corner_points[0][1])
        grade_br_row = int(grade_corner_points[3][0])
        grade_br_col = int(grade_corner_points[3][1])
        grade_wd = grade_br_col - grade_tl_col
        grade_ht = grade_br_row - grade_tl_row
    
    # For array slicing: imgGray[rows, cols]
    y, x, h, w = grade_tl_row, grade_tl_col, grade_ht, grade_wd
    grade_roi = imgGray[y:y + h, x:x + w]
    
    if grade_roi.size > 0:
        grade_new_image = np.zeros_like(grade_roi)
        grade_new_image[:, :] = grade_roi
    else:
        print(f"⚠️ Warning: Grade ROI extraction failed, skipping grade display")
        grade_new_image = None
    
    helper.manual_draw_contours(grade_contour, contours[grade_index], (0, 255, 0), 1)
    cv2.imshow('Grading section', grade_contour)
    cv2.waitKey(0)

    imgGrading = grade_contour.copy()
    cv2.putText(imgGrading, str(int(score)) + "%", 
                config.GRADE_TEXT_POSITION, 
                config.GRADE_TEXT_FONT, 
                config.GRADE_TEXT_SCALE, 
                config.GRADE_TEXT_COLOR, 
                config.GRADE_TEXT_THICKNESS)
    cv2.imshow("Grading", imgGrading)
    cv2.waitKey(0)

    cv2.destroyAllWindows()
    
    
    

def show_result_summary(id_list, score_list):
    """Display comprehensive result summary in console"""
    
    print("\n" + "="*70)
    print(" "*20 + "📊 RESULT SUMMARY 📊")
    print("="*70)
    
    total_students = len(score_list)
    
    if total_students == 0:
        print("No results to display.")
        return
    
    # Calculate statistics
    average_score = sum(score_list) / total_students
    highest_score = max(score_list)
    lowest_score = min(score_list)
    
    # Pass/Fail analysis
    pass_mark = config.PASS_MARK
    passed_students = sum(1 for score in score_list if score >= pass_mark)
    failed_students = total_students - passed_students
    pass_rate = (passed_students / total_students) * 100
    
    # Grade distribution
    grade_a = sum(1 for score in score_list if score >= config.GRADE_BOUNDARIES['A'])
    grade_b = sum(1 for score in score_list if config.GRADE_BOUNDARIES['B'] <= score < config.GRADE_BOUNDARIES['A'])
    grade_c = sum(1 for score in score_list if config.GRADE_BOUNDARIES['C'] <= score < config.GRADE_BOUNDARIES['B'])
    grade_f = sum(1 for score in score_list if score < config.GRADE_BOUNDARIES['C'])
    
    # Display overall statistics
    print(f"\n📈 OVERALL STATISTICS:")
    print(f"   Total Students     : {total_students}")
    print(f"   Average Score      : {average_score:.2f}%")
    print(f"   Highest Score      : {highest_score:.2f}%")
    print(f"   Lowest Score       : {lowest_score:.2f}%")
    print(f"   Pass Rate          : {pass_rate:.2f}% ({passed_students}/{total_students})")
    
    print(f"\n📊 GRADE DISTRIBUTION:")
    print(f"   Grade A (80-100%)  : {grade_a} students ({(grade_a/total_students)*100:.1f}%)")
    print(f"   Grade B (60-79%)   : {grade_b} students ({(grade_b/total_students)*100:.1f}%)")
    print(f"   Grade C (40-59%)   : {grade_c} students ({(grade_c/total_students)*100:.1f}%)")
    print(f"   Grade F (0-39%)    : {grade_f} students ({(grade_f/total_students)*100:.1f}%)")
    
    # Display individual results
    print(f"\n📝 INDIVIDUAL RESULTS:")
    print("-"*70)
    print(f"{'No.':<6} {'Student ID':<25} {'Score':<12} {'Grade':<10} {'Status'}")
    print("-"*70)
    
    for i, (student_id, score) in enumerate(zip(id_list, score_list), 1):
        # Determine grade
        if score >= config.GRADE_BOUNDARIES['A']:
            grade = "A"
        elif score >= config.GRADE_BOUNDARIES['B']:
            grade = "B"
        elif score >= config.GRADE_BOUNDARIES['C']:
            grade = "C"
        else:
            grade = "F"
        
        # Determine status
        status = "✓ PASS" if score >= pass_mark else "✗ FAIL"
        
        print(f"{i:<6} {student_id:<25} {score:>6.2f}%     {grade:<10} {status}")
    
    print("-"*70)
    
    # Performance insights
    print(f"\n💡 INSIGHTS:")
    if pass_rate >= 80:
        print(f"   🌟 Excellent performance! {pass_rate:.1f}% pass rate.")
    elif pass_rate >= 60:
        print(f"   ✓ Good performance with {pass_rate:.1f}% pass rate.")
    elif pass_rate >= 40:
        print(f"   ⚠ Average performance. {pass_rate:.1f}% pass rate - needs improvement.")
    else:
        print(f"   ⚠ Poor performance. Only {pass_rate:.1f}% pass rate - significant improvement needed.")
    
    if highest_score == 100:
        print(f"   🏆 Perfect score achieved!")
    
    if average_score < 50:
        print(f"   📚 Class average is below 50%. Consider reviewing difficult topics.")
    
    print("\n" + "="*70 + "\n")

def print_result():
    global uploaded_images
    if uploaded_images:

        # result_text = "Uploaded images:\n"
        for file_path, _ in uploaded_images:
            print(file_path)
            processImages(file_path)

    else:
        result_label.config(text="No images uploaded yet.")
    
    # Display summary statistics if multiple students were processed
    if len(id) > 0:
        show_result_summary(id, res)
    
    # Save to Excel
    save_to_excel(id,res)

    if print_result:
        result_text = "Results saved successfully!"
        result_label.config(text=result_text)



def save_to_excel(id_array, score_array):
    wb = Workbook()
    ws = wb.active
    ws.title = "ID Scores"

    # Write headers
    ws['A1'] = "No."
    ws['B1'] = "Student ID"
    ws['C1'] = "Score (%)"
    ws['D1'] = "Grade"
    ws['E1'] = "Status"

    # Calculate statistics for summary
    total_students = len(score_array)
    pass_mark = config.PASS_MARK
    
    # Write individual student data
    for idx, (id_val, score_val) in enumerate(zip(id_array, score_array), start=2):
        # Determine grade
        if score_val >= config.GRADE_BOUNDARIES['A']:
            grade = "A"
        elif score_val >= config.GRADE_BOUNDARIES['B']:
            grade = "B"
        elif score_val >= config.GRADE_BOUNDARIES['C']:
            grade = "C"
        else:
            grade = "F"
        
        # Determine status
        status = "PASS" if score_val >= pass_mark else "FAIL"
        
        ws[f'A{idx}'] = idx - 1  # Student number
        ws[f'B{idx}'] = id_val
        ws[f'C{idx}'] = round(score_val, 2)
        ws[f'D{idx}'] = grade
        ws[f'E{idx}'] = status
    
    # Add summary statistics at the bottom
    summary_row = len(score_array) + 3
    
    ws[f'A{summary_row}'] = "SUMMARY STATISTICS"
    ws[f'A{summary_row}'].font = Font(bold=True)
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Total Students:"
    ws[f'B{summary_row}'] = total_students
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Average Score:"
    ws[f'B{summary_row}'] = round(sum(score_array) / total_students, 2) if total_students > 0 else 0
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Highest Score:"
    ws[f'B{summary_row}'] = round(max(score_array), 2) if score_array else 0
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Lowest Score:"
    ws[f'B{summary_row}'] = round(min(score_array), 2) if score_array else 0
    
    summary_row += 1
    passed_students = sum(1 for score in score_array if score >= pass_mark)
    ws[f'A{summary_row}'] = "Pass Rate:"
    ws[f'B{summary_row}'] = f"{(passed_students / total_students * 100):.2f}%" if total_students > 0 else "0%"
    
    summary_row += 2
    ws[f'A{summary_row}'] = "GRADE DISTRIBUTION"
    ws[f'A{summary_row}'].font = Font(bold=True)
    
    summary_row += 1
    grade_a = sum(1 for score in score_array if score >= config.GRADE_BOUNDARIES['A'])
    ws[f'A{summary_row}'] = f"Grade A ({config.GRADE_BOUNDARIES['A']}-100%):"
    ws[f'B{summary_row}'] = grade_a
    ws[f'C{summary_row}'] = f"{(grade_a/total_students)*100:.1f}%" if total_students > 0 else "0%"
    
    summary_row += 1
    grade_b = sum(1 for score in score_array if config.GRADE_BOUNDARIES['B'] <= score < config.GRADE_BOUNDARIES['A'])
    ws[f'A{summary_row}'] = f"Grade B ({config.GRADE_BOUNDARIES['B']}-{config.GRADE_BOUNDARIES['A']-1}%):"
    ws[f'B{summary_row}'] = grade_b
    ws[f'C{summary_row}'] = f"{(grade_b/total_students)*100:.1f}%" if total_students > 0 else "0%"
    
    summary_row += 1
    grade_c = sum(1 for score in score_array if config.GRADE_BOUNDARIES['C'] <= score < config.GRADE_BOUNDARIES['B'])
    ws[f'A{summary_row}'] = f"Grade C ({config.GRADE_BOUNDARIES['C']}-{config.GRADE_BOUNDARIES['B']-1}%):"
    ws[f'B{summary_row}'] = grade_c
    ws[f'C{summary_row}'] = f"{(grade_c/total_students)*100:.1f}%" if total_students > 0 else "0%"
    
    summary_row += 1
    grade_f = sum(1 for score in score_array if score < config.GRADE_BOUNDARIES['C'])
    ws[f'A{summary_row}'] = f"Grade F (0-{config.GRADE_BOUNDARIES['C']-1}%):"
    ws[f'B{summary_row}'] = grade_f
    ws[f'C{summary_row}'] = f"{(grade_f/total_students)*100:.1f}%" if total_students > 0 else "0%"

    # Adjust column widths for better readability
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10

    # Save the workbook with safe filename
    from datetime import datetime
    
    # Generate a safe default filename with timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    default_filename = f"OMR_Results_{timestamp}.xlsx"
    
    try:
        excel_filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Save OMR Results",
            initialfile=default_filename
        )
        
        if excel_filename:
            # Ensure the filename ends with .xlsx
            if not excel_filename.endswith('.xlsx'):
                excel_filename += '.xlsx'
            
            wb.save(excel_filename)
            print(f"✅ Excel file saved successfully to: {excel_filename}")
        else:
            print("⚠️ Save cancelled by user.")
    except Exception as e:
        print(f"❌ Error saving Excel file: {e}")
        # Try to save to desktop as fallback
        try:
            desktop = os.path.join(os.path.expanduser("~"), "Desktop")
            fallback_path = os.path.join(desktop, default_filename)
            wb.save(fallback_path)
            print(f"✅ File saved to desktop: {fallback_path}")
        except Exception as e2:
            print(f"❌ Failed to save file: {e2}")

def rounded_button(parent, text, command, width=20, bg_color="#4299e1"):
    button = tk.Button(parent, text=text, command=command, bg=bg_color, fg="white", 
                       padx=20, pady=10, relief="flat", font=("Arial", 12, "bold"), 
                       width=width, cursor="hand2",
                       activebackground="#3182ce", activeforeground="white")
    button.config(borderwidth=0, highlightthickness=0, bd=0)
    return button

def update_config():
    """Update config values from UI inputs"""
    try:
        num_q = int(questions_entry.get())
        num_c = int(choices_entry.get())
        
        if num_q < 1 or num_q > 100:
            result_label.config(text="❌ Number of questions must be between 1 and 100!", 
                              fg="#fc8181")
            return False
        
        if num_c < 2 or num_c > 10:
            result_label.config(text="❌ Number of choices must be between 2 and 10!", 
                              fg="#fc8181")
            return False
        
        # Update config
        config.NUM_QUESTIONS = num_q
        config.NUM_CHOICES = num_c
        
        # Generate default answer key (all zeros)
        config.ANSWER_KEY = [0] * num_q
        
        # Update answer key display
        generate_answer_inputs()
        
        result_label.config(text=f"✅ Configuration updated: {num_q} questions with {num_c} choices each\n📝 Now set the correct answers below!", 
                          fg="#68d391")
        return True
        
    except ValueError:
        result_label.config(text="❌ Please enter valid numbers!", fg="#fc8181")
        return False

def generate_answer_inputs():
    """Generate input fields for answer key based on number of questions"""
    # Clear existing widgets
    for widget in answer_frame.winfo_children():
        widget.destroy()
    
    num_q = config.NUM_QUESTIONS
    num_c = config.NUM_CHOICES
    
    # Title with instruction
    answer_title = tk.Label(answer_frame, 
                           text=f"📋 Answer Key", 
                           font=("Arial", 13, "bold"), 
                           bg="#2d3748", fg="#f7fafc")
    answer_title.pack(pady=(15, 5))
    
    instruction = tk.Label(answer_frame, 
                          text=f"Enter the correct answer (0={chr(65)} to {num_c-1}={chr(65+num_c-1)}) for each question:", 
                          font=("Arial", 9), 
                          bg="#2d3748", fg="#cbd5e0")
    instruction.pack(pady=(0, 10))
    
    # Create scrollable frame for answers
    answer_canvas = tk.Canvas(answer_frame, bg="#2d3748", height=200, highlightthickness=0)
    answer_scrollbar = tk.Scrollbar(answer_frame, orient="vertical", command=answer_canvas.yview)
    scrollable_frame = tk.Frame(answer_canvas, bg="#2d3748")
    
    scrollable_frame.bind(
        "<Configure>",
        lambda e: answer_canvas.configure(scrollregion=answer_canvas.bbox("all"))
    )
    
    answer_canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
    answer_canvas.configure(yscrollcommand=answer_scrollbar.set)
    
    # Create answer entry fields in a grid
    global answer_entries
    answer_entries = []
    
    cols = 5  # 5 questions per row for better visibility
    for i in range(num_q):
        row = i // cols
        col = i % cols
        
        # Question card
        q_frame = tk.Frame(scrollable_frame, bg="#4a5568", relief="flat", bd=1, 
                          highlightbackground="#718096", highlightthickness=1)
        q_frame.grid(row=row, column=col, padx=8, pady=8, sticky="nsew")
        
        q_label = tk.Label(q_frame, text=f"Q {i+1}", 
                          font=("Arial", 10, "bold"), 
                          bg="#4a5568", fg="#f7fafc", width=6)
        q_label.pack(pady=(8, 2))
        
        q_entry = tk.Entry(q_frame, font=("Arial", 12, "bold"), width=4, 
                          justify="center", bg="#edf2f7", fg="#2d3748",
                          relief="flat", bd=2)
        q_entry.insert(0, str(config.ANSWER_KEY[i]))
        q_entry.pack(pady=(2, 8), padx=8)
        
        answer_entries.append(q_entry)
    
    # Configure grid weights for proper spacing
    for col in range(cols):
        scrollable_frame.grid_columnconfigure(col, weight=1, uniform="equal")
    
    answer_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=10, pady=5)
    answer_scrollbar.pack(side=tk.RIGHT, fill=tk.Y, pady=5)
    
    # Bind mouse wheel to canvas
    def _on_mousewheel(event):
        answer_canvas.yview_scroll(int(-1*(event.delta/120)), "units")
    answer_canvas.bind_all("<MouseWheel>", _on_mousewheel)
    
    # Save answer key button
    save_answers_btn = tk.Button(answer_frame, text="✓ Save Answer Key", command=save_answer_key,
                                 bg="#48bb78", fg="white", font=("Arial", 11, "bold"),
                                 padx=30, pady=8, relief="flat", cursor="hand2",
                                 activebackground="#38a169", activeforeground="white")
    save_answers_btn.pack(pady=(10, 15))

def save_answer_key():
    """Save the answer key from input fields"""
    try:
        new_answers = []
        for i, entry in enumerate(answer_entries):
            val = int(entry.get())
            if val < 0 or val >= config.NUM_CHOICES:
                result_label.config(text=f"❌ Question {i+1}: Answer must be between 0 and {config.NUM_CHOICES-1}!", 
                                  fg="#fc8181")
                return False
            new_answers.append(val)
        
        config.ANSWER_KEY = new_answers
        
        # Convert to letter format for display
        letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']
        answer_display = ', '.join([f"Q{i+1}:{letters[ans]}" for i, ans in enumerate(new_answers[:10])])
        if len(new_answers) > 10:
            answer_display += f" ... +{len(new_answers)-10} more"
        
        result_label.config(text=f"✅ Answer Key Saved Successfully!\n{answer_display}", 
                          fg="#68d391")
        
        # Open the upload/process popup
        open_processing_popup()
        
        return True
        
    except ValueError:
        result_label.config(text="❌ Please enter valid numbers (0 to " + str(config.NUM_CHOICES-1) + ") for all answers!", 
                          fg="#fc8181")
        return False

def open_processing_popup():
    """Open a popup window for image upload and processing"""
    popup = tk.Toplevel(root)
    popup.title("OMR Processing Center")
    popup.geometry("700x600")
    popup.configure(bg="#1a202c")
    popup.resizable(True, True)
    
    # Make popup modal
    popup.transient(root)
    popup.grab_set()
    
    # Center the popup
    popup.update_idletasks()
    x = (popup.winfo_screenwidth() // 2) - (700 // 2)
    y = (popup.winfo_screenheight() // 2) - (600 // 2)
    popup.geometry(f"700x600+{x}+{y}")
    
    # Main container
    popup_container = tk.Frame(popup, bg="#1a202c")
    popup_container.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
    
    # Header
    header = tk.Frame(popup_container, bg="#2d3748")
    header.pack(fill=tk.X, pady=(0, 20))
    
    title = tk.Label(header, text="🎯 OMR Processing Center", 
                    font=("Arial", 18, "bold"), 
                    bg="#2d3748", fg="#f7fafc")
    title.pack(pady=15)
    
    # Configuration summary
    summary_frame = tk.Frame(popup_container, bg="#2d3748", relief="flat")
    summary_frame.pack(fill=tk.X, pady=(0, 15))
    
    summary_title = tk.Label(summary_frame, text="📊 Current Configuration", 
                            font=("Arial", 12, "bold"), 
                            bg="#2d3748", fg="#f7fafc")
    summary_title.pack(pady=(10, 5))
    
    config_text = f"Questions: {config.NUM_QUESTIONS}  |  Choices: {config.NUM_CHOICES}  |  Answer Key: Set ✓"
    config_label = tk.Label(summary_frame, text=config_text, 
                           font=("Arial", 10), 
                           bg="#2d3748", fg="#cbd5e0")
    config_label.pack(pady=(0, 10))
    
    # Upload section
    upload_section = tk.Frame(popup_container, bg="#2d3748")
    upload_section.pack(fill=tk.BOTH, expand=True, pady=(0, 15))
    
    upload_title = tk.Label(upload_section, text="📁 Upload Answer Sheets", 
                           font=("Arial", 13, "bold"), 
                           bg="#2d3748", fg="#f7fafc")
    upload_title.pack(pady=(15, 10))
    
    upload_status = tk.Label(upload_section, text="No images uploaded yet", 
                            font=("Arial", 10), 
                            bg="#2d3748", fg="#cbd5e0")
    upload_status.pack(pady=5)
    
    # Image preview canvas
    preview_canvas = tk.Canvas(upload_section, bg="#1a202c", height=200, highlightthickness=1,
                              highlightbackground="#4a5568")
    preview_canvas.pack(fill=tk.BOTH, expand=True, padx=15, pady=10)
    
    preview_scrollbar = tk.Scrollbar(upload_section, orient=tk.HORIZONTAL, command=preview_canvas.xview)
    preview_canvas.configure(xscrollcommand=preview_scrollbar.set)
    preview_scrollbar.pack(fill=tk.X, padx=15)
    
    preview_frame = tk.Frame(preview_canvas, bg="#1a202c")
    preview_canvas.create_window((0, 0), window=preview_frame, anchor="nw")
    
    def update_preview():
        # Check if widgets still exist before updating
        try:
            if not preview_frame.winfo_exists():
                return
        except tk.TclError:
            return
        
        # Clear previous thumbnails
        for widget in preview_frame.winfo_children():
            widget.destroy()
        
        # Display thumbnails
        for img_tk in thumbnail_images:
            img_container = tk.Frame(preview_frame, bg="#4a5568", relief="flat",
                                    highlightbackground="#718096", highlightthickness=1)
            img_container.pack(side=tk.LEFT, padx=8, pady=8)
            
            label = tk.Label(img_container, image=img_tk, bg="#4a5568")
            label.pack(padx=5, pady=5)
        
        preview_frame.update_idletasks()
        preview_canvas.configure(scrollregion=preview_canvas.bbox("all"))
        
        if uploaded_images:
            upload_status.config(text=f"✅ {len(uploaded_images)} image(s) uploaded and ready for processing",
                               fg="#68d391")
    
    def popup_upload():
        global uploaded_images, thumbnail_images
        file_paths = filedialog.askopenfilenames(
            filetypes=[("Image Files", "*.png;*.jpg;*.jpeg;*.gif;*.bmp")],
            title="Choose answer sheets to process"
        )
        for file_path in file_paths:
            image = Image.open(file_path)
            image.thumbnail(config.THUMBNAIL_SIZE)
            uploaded_images.append((file_path, image))
            thumbnail_images.append(ImageTk.PhotoImage(image))
        
        update_preview()
    
    def popup_process():
        if not uploaded_images:
            upload_status.config(text="❌ Please upload images first!", fg="#fc8181")
            return
        
        popup.destroy()
        print_result()
    
    # Upload button
    upload_btn = tk.Button(upload_section, text="📁 Select Answer Sheets", 
                          command=popup_upload,
                          bg="#805ad5", fg="white", font=("Arial", 11, "bold"),
                          padx=30, pady=10, relief="flat", cursor="hand2",
                          activebackground="#6b46c1", activeforeground="white")
    upload_btn.pack(pady=10)
    
    # Action buttons
    button_frame = tk.Frame(popup_container, bg="#1a202c")
    button_frame.pack(fill=tk.X, pady=(10, 0))
    
    process_btn = tk.Button(button_frame, text="🚀 Process All Images", 
                           command=popup_process,
                           bg="#38b2ac", fg="white", font=("Arial", 12, "bold"),
                           padx=40, pady=12, relief="flat", cursor="hand2",
                           activebackground="#319795", activeforeground="white")
    process_btn.pack(side=tk.LEFT, expand=True, padx=5)
    
    cancel_btn = tk.Button(button_frame, text="❌ Cancel", 
                          command=popup.destroy,
                          bg="#718096", fg="white", font=("Arial", 12, "bold"),
                          padx=40, pady=12, relief="flat", cursor="hand2",
                          activebackground="#4a5568", activeforeground="white")
    cancel_btn.pack(side=tk.LEFT, expand=True, padx=5)


def upload_images():
    """Legacy function - now handled by popup"""
    pass


if __name__ == "__main__":
    uploaded_images = []
    thumbnail_images = []

    root = tk.Tk()
    root.title("OMR Answer Sheet Scanner - Professional Edition")

    root.geometry("850x750")  # Optimized size
    root.configure(bg="#1a202c")  # Dark professional background

    # Create main container with padding
    main_container = tk.Frame(root, bg="#1a202c")
    main_container.pack(fill=tk.BOTH, expand=True, padx=25, pady=20)

    # ==== HEADER ====
    header_frame = tk.Frame(main_container, bg="#2d3748", relief="flat")
    header_frame.pack(fill=tk.X, pady=(0, 20))
    
    title_label = tk.Label(header_frame, text="📝 OMR Answer Sheet Scanner", 
                          font=("Arial", 20, "bold"), 
                          bg="#2d3748", fg="#f7fafc")
    title_label.pack(pady=15)
    
    subtitle_label = tk.Label(header_frame, text="Professional Optical Mark Recognition System", 
                             font=("Arial", 10), 
                             bg="#2d3748", fg="#cbd5e0")
    subtitle_label.pack(pady=(0, 15))

    # ==== CONFIGURATION SECTION ====
    config_frame = tk.Frame(main_container, bg="#2d3748", relief="flat", bd=0)
    config_frame.pack(fill=tk.X, pady=(0, 15))

    config_title = tk.Label(config_frame, text="⚙️ Exam Configuration", 
                           font=("Arial", 14, "bold"), 
                           bg="#2d3748", fg="#f7fafc")
    config_title.pack(pady=(15, 10))

    # Input grid for questions and choices
    input_frame = tk.Frame(config_frame, bg="#2d3748")
    input_frame.pack(pady=(5, 10))

    # Number of Questions
    questions_label = tk.Label(input_frame, text="Number of Questions:", 
                              font=("Arial", 11, "bold"), bg="#2d3748", fg="#e2e8f0")
    questions_label.grid(row=0, column=0, padx=15, pady=8, sticky="e")

    questions_entry = tk.Entry(input_frame, font=("Arial", 12), width=12, justify="center",
                              bg="#edf2f7", fg="#2d3748", relief="flat", bd=2)
    questions_entry.insert(0, str(config.NUM_QUESTIONS))
    questions_entry.grid(row=0, column=1, padx=15, pady=8)

    # Number of Choices
    choices_label = tk.Label(input_frame, text="Choices per Question:", 
                            font=("Arial", 11, "bold"), bg="#2d3748", fg="#e2e8f0")
    choices_label.grid(row=1, column=0, padx=15, pady=8, sticky="e")

    choices_entry = tk.Entry(input_frame, font=("Arial", 12), width=12, justify="center",
                            bg="#edf2f7", fg="#2d3748", relief="flat", bd=2)
    choices_entry.insert(0, str(config.NUM_CHOICES))
    choices_entry.grid(row=1, column=1, padx=15, pady=8)

    # Update button
    update_btn = tk.Button(config_frame, text="✓ Apply Configuration", command=update_config,
                          bg="#4299e1", fg="white", font=("Arial", 11, "bold"),
                          padx=30, pady=8, relief="flat", cursor="hand2",
                          activebackground="#3182ce", activeforeground="white")
    update_btn.pack(pady=(5, 15))

    # ==== ANSWER KEY SECTION ====
    answer_frame = tk.Frame(main_container, bg="#2d3748", relief="flat", bd=0)
    answer_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 15))
    
    # Initialize with current answer key
    answer_entries = []
    generate_answer_inputs()

    # ==== STATUS LABEL ====
    result_label = tk.Label(main_container, 
                           text="👆 Configure the exam settings above and set your answer key to begin", 
                           wraplength=750, bg="#1a202c", padx=15, pady=15,
                           font=("Arial", 11), fg="#cbd5e0",
                           relief="flat", bd=1, highlightbackground="#4a5568", 
                           highlightthickness=1)
    result_label.pack(pady=(5, 15), fill=tk.X)

    # ==== INFO SECTION ====
    info_frame = tk.Frame(main_container, bg="#2d3748", relief="flat")
    info_frame.pack(fill=tk.BOTH, expand=True, pady=(10, 0))
    
    info_title = tk.Label(info_frame, text="ℹ️ How to Use", 
                         font=("Arial", 12, "bold"), 
                         bg="#2d3748", fg="#f7fafc")
    info_title.pack(pady=(15, 10))
    
    instructions = [
        "1️⃣ Set the number of questions and choices",
        "2️⃣ Click 'Apply Configuration' to generate answer fields",
        "3️⃣ Enter the correct answer for each question (0=A, 1=B, etc.)",
        "4️⃣ Click 'Save Answer Key' to open the processing window",
        "5️⃣ Upload your scanned answer sheets and process them!"
    ]
    
    for instruction in instructions:
        inst_label = tk.Label(info_frame, text=instruction, 
                             font=("Arial", 10), bg="#2d3748", fg="#cbd5e0",
                             anchor="w")
        inst_label.pack(pady=2, padx=20, anchor="w")

    root.mainloop()